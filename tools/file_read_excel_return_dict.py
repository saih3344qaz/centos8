#!/usr/bin/env python3
# -*- coding=utf-8 -*-

from openpyxl import Workbook
from openpyxl import load_workbook


def excel_parser_return_dict(file='test.xlsx', sheel_name='Sheet1'):
    data = load_workbook(file)  # 读取xlsx文件
    table = data[sheel_name]  # 读取sheet数据
    excel_dict = {}
    row_location = 0
    for row in table.iter_rows():
        if row_location == 0:  # 跳过第一行！
            row_location += 1
            continue
        else:
            cell_location = 0
            for cell in row:
                if cell_location == 0:  # 读取第一列的用户名
                    username = cell.value
                    cell_location += 1
                elif cell_location == 1:  # 读取第二列的密码
                    password = cell.value
                    cell_location += 1
                elif cell_location == 2:  # 读取第三列的级别
                    privilege = cell.value
                    cell_location += 1
            excel_dict[username] = password, privilege  # 写入字典
        row_location += 1
    return excel_dict  # 返回字典


def excel_parser_return_list(file='test.xlsx', sheel_name='Sheet1'):
    data = load_workbook(file)  # 读取xlsx文件
    table = data[sheel_name]  # 读取sheet数据
    excel_list = []
    row_location = 0
    for row in table.iter_rows():
        if row_location == 0:  # 跳过第一行！
            row_location += 1
            continue
        else:
            cell_location = 0
            for cell in row:
                if cell_location == 0:  # 读取第一列的用户名
                    username = cell.value
                    cell_location += 1
                elif cell_location == 1:  # 读取第二列的密码
                    password = cell.value
                    cell_location += 1
                elif cell_location == 2:  # 读取第三列的级别
                    privilege = cell.value
                    cell_location += 1
            excel_list.append({'username': username, 'password': password, 'privilege': privilege})  # 写入字典
        row_location += 1
    return excel_list  # 返回列表


if __name__ == "__main__":
    # print(excel_parser_return_dict('./excel_file/read_accounts.xlsx'))
    print(excel_parser_return_list('./excel_file/read_accounts.xlsx'))

