#!/usr/bin/env python3
# -*- coding=utf-8 -*-

import openpyxl

# 读取Excel文件并格式化，将各个设备的信息存入字典。
def read_excel(file_name):
    wb = openpyxl.load_workbook(file_name)
    sheet = wb.get_sheet_by_name('Sheet1')
    row = sheet.max_row
    column = sheet.max_column
    device_list = {}
    for i in range(2, row + 1):
        device_list['device{0}'.format(i - 1)] = []
        for j in range(1, column + 1):
            vla = sheet.cell(row=i, column=j).value
            device_list['device{0}'.format(i - 1)].append(vla)
    return device_list


if __name__ == '__main__':
    read_excel()
