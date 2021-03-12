#!/usr/bin/env python3
# -*- coding=utf-8 -*-

from openpyxl import Workbook
from openpyxl import load_workbook

dict_excel = {'test123': ('cisco123', 15), 'test456': ('cisco456', 1), 'test789': ('cisco789', 1)}
dict_list = [{'username': 'test123', 'privilege': 15, 'password': 'Cisc0123'},
             {'username': 'test124', 'privilege': 1, 'password': 'Cisc0123'}]


def excel_write(file='./excel_file/write_pyxl.xlsx', sheel_name='Sheet1', write_dict=dict_excel):
    wb = Workbook()  # 创建xlsx
    ws = wb.create_sheet()  # 创建sheet
    ws.title = sheel_name  # 命名sheet
    # 写入第一行内容
    ws['A1'] = '用户'
    ws['B1'] = '密码'
    ws['C1'] = '级别'
    row_location = 2  # 从第二行开始写入内容
    for x, y in write_dict.items():
        user_locatin = 'A' + str(row_location)
        pass_locatin = 'B' + str(row_location)
        priv_locatin = 'C' + str(row_location)
        ws[user_locatin] = x  # 写入用户
        ws[pass_locatin] = y[0]  # 写入密码
        ws[priv_locatin] = y[1]  # 写入级别
        row_location += 1  # 行号加1
    wb.save(file)  # 保存xlsx文件


def excel_write_list(file='./excel_file/write_pyxl.xlsx', sheel_name='Sheet1', write_list=dict_list):
    wb = Workbook()  # 创建xlsx
    ws = wb.create_sheet()  # 创建sheet
    ws.title = sheel_name  # 命名sheet
    # 写入第一行内容
    ws['A1'] = '用户'
    ws['B1'] = '密码'
    ws['C1'] = '级别'
    row_location = 2  # 从第二行开始写入内容
    for user_dict in write_list:
        user_locatin = f'A{row_location}'
        pass_locatin = f'B{row_location}'
        priv_locatin = f'C{row_location}'
        ws[user_locatin] = user_dict.get('username')
        ws[pass_locatin] = user_dict.get('password')
        ws[priv_locatin] = user_dict.get('privilege')
        row_location += 1  # 行号加1
    wb.save(file)  # 保存xlsx文件


if __name__ == "__main__":
    # excel_write()
    excel_write_list()

