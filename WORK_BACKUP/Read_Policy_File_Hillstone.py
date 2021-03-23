#!/usr/bin/python3.9
# -*- coding=utf-8 -*-
# 本脚由孙艳龙编写，用于Python学习及网络自动化脚本，如有问题或技术交流请与本人联系！
# mail:sunyanlong@bris.cn

import openpyxl
import re


def read_excel(file_name):
    wb = openpyxl.load_workbook(file_name)
    sheet = wb.get_sheet_by_name('Sheet1')
    row = sheet.max_row
    column = sheet.max_column
    policy_list = {}
    for i in range(2, row + 1):
        policy_list['policy{0}'.format(i - 1)] = []
        for j in range(1, column + 1):
            vla = sheet.cell(row=i, column=j).value
            policy_list['policy{0}'.format(i - 1)].append(vla)
    return policy_list


def create_sadd(excelfile):  # 创建源地址
    sadd_dict = read_excel(excelfile)
    sadd_cmd = []
    sadd_cmds = []
    for x in sadd_dict.values():
        x2 = x[2]
        src_list = x2.split(' ')
        for y in src_list:
            if re.search(r'(\d{1,3}\.\d{1,3}\.\d{1,3}\.\d)[/][2][4]', y):
                sadd_1 = 'address ' + str(y)
                sadd_2 = 'ip ' + str(y)
                sadd_3 = 'exit'
                sadd_cmd.extend((sadd_1, sadd_2, sadd_3))
                sadd_cmds = '\n'.join(sadd_cmd)
            elif re.search(r'\d{1,3}\.\d{1,3}\.[0]\.[0][/][1][6]', y):
                sadd_1 = 'address ' + str(y)
                sadd_2 = 'ip ' + str(y)
                sadd_3 = 'exit'
                sadd_cmd.extend((sadd_1, sadd_2, sadd_3))
                sadd_cmds = '\n'.join(sadd_cmd)
            elif re.search(r'\d{1,3}\.\d{1,3}\.\d{1,3}\.\d{1,3}[-]\d{1,3}', y):
                range_list = re.split(r'[-.]', y)
                range_list1 = re.split(r'[-]', y)
                sadd_1 = 'address ' + str(range_list[0]) + '.' + str(range_list[1]) + '.' + str(
                    range_list[2]) + '.' + '[' + str(range_list[3]) + '-' + str(range_list[4]) + ']'
                sadd_2 = 'range ' + str(range_list1[0]) + ' ' + str(range_list[0]) + '.' + str(
                    range_list[1]) + '.' + str(range_list[2]) + '.' + str(range_list[4])
                sadd_3 = 'exit'
                sadd_cmd.extend((sadd_1, sadd_2, sadd_3))
                sadd_cmds = '\n'.join(sadd_cmd)
            else:
                sadd_1 = 'address ' + str(y) + '/32'
                sadd_2 = 'ip ' + str(y) + '/32'
                sadd_3 = 'exit'
                sadd_cmd.extend((sadd_1, sadd_2, sadd_3))
                sadd_cmds = '\n'.join(sadd_cmd)
    return sadd_cmds


def create_dadd(excelfile):  # 创建目的地址
    dadd_dice = read_excel(excelfile)
    dadd_cmd = []
    dadd_cmds = []
    for x in dadd_dice.values():
        x4 = x[4]
        dst_list = x4.split(' ')
        for y in dst_list:
            if re.search(r'(\d{1,3}\.\d{1,3}\.\d{1,3}\.\d)[/][2][4]', y):
                dadd_name = 'address ' + str(y)
                dadd_add = 'ip ' + str(y)
                dadd_exit = 'exit'
                dadd_cmd.extend((dadd_name, dadd_add, dadd_exit))
                dadd_cmds = '\n'.join(dadd_cmd)
            elif re.search(r'\d{1,3}\.\d{1,3}\.[0]\.[0][/][1][6]', y):
                dadd_1 = 'address ' + str(y)
                dadd_2 = 'ip ' + str(y)
                dadd_3 = 'exit'
                dadd_cmd.extend((dadd_1, dadd_2, dadd_3))
                dadd_cmds = '\n'.join(dadd_cmd)
            elif re.search(r'\d{1,3}\.\d{1,3}\.\d{1,3}\.\d{1,3}[-]\d{1,3}', y):
                range_list = re.split(r'[-.]', y)
                range_list1 = re.split(r'[-]', y)
                dadd_1 = 'address ' + str(range_list[0]) + '.' + str(range_list[1]) + '.' + str(
                    range_list[2]) + '.' + '[' + str(range_list[3]) + '-' + str(range_list[4]) + ']'
                dadd_2 = 'range ' + str(range_list1[0]) + ' ' + str(range_list[0]) + '.' + str(
                    range_list[1]) + '.' + str(range_list[2]) + '.' + str(range_list[4])
                dadd_3 = 'exit'
                dadd_cmd.extend((dadd_1, dadd_2, dadd_3))
                dadd_cmds = '\n'.join(dadd_cmd)
            else:
                dadd_1 = 'address ' + str(y) + '/32'
                dadd_2 = 'ip ' + str(y) + '/32'
                dadd_3 = 'exit'
                dadd_cmd.extend((dadd_1, dadd_2, dadd_3))
                dadd_cmds = '\n'.join(dadd_cmd)
    return dadd_cmds


def create_service(excelfile):  # 创建service
    ser_dict = read_excel(excelfile)
    ser_cmd = []
    ser_cmds = []
    for x in ser_dict.values():
        x6 = x[6]
        x6_str = str(x6)
        ser_list = x6_str.split(' ')
        for y in ser_list:
            if re.search(r'\d{1,5}[-]\d{1,5}', y):
                range_ser = re.split(r'[-]', y)
                ser_name = 'service ' + str(x[5]).upper() + str(y)
                ser_port = str(x[5]) + ' dst-port ' + \
                    str(range_ser[0]) + ' ' + str(range_ser[1])
                ser_exit = 'exit'
                ser_cmd.extend((ser_name, ser_port, ser_exit))
                ser_cmds = '\n'.join(ser_cmd)
            else:
                ser_name = 'service ' + str(x[5]).upper() + str(y)
                ser_port = str(x[5]) + ' dst-port ' + str(y)
                ser_exit = 'exit'
                ser_cmd.extend((ser_name, ser_port, ser_exit))
                ser_cmds = '\n'.join(ser_cmd)
    return ser_cmds


def create_policy(excelfile):  # 创建策略
    policy_dict = read_excel(excelfile)
    policy_cmd = []
    policy_cmds = []
    for x in policy_dict.values():
        rule = 'rule'
        act = 'action permit'
        szone = 'src-zone ' + str(x[1])
        dzone = 'dst-zone ' + str(x[3])
        srcadd = []
        dstadd = []
        serv = []
        exit = 'exit'
        x2 = x[2]
        src_list = x2.split(' ')
        for y in src_list:
            if re.search(r'(\d{1,3}\.\d{1,3}\.\d{1,3}\.\d)[/][2][4]', y):
                y1 = 'src-addr ' + y
                srcadd.append(y1)
            elif re.search(r'\d{1,3}\.\d{1,3}\.[0]\.[0][/][1][6]', y):
                y1 = 'src-addr ' + y
                srcadd.append(y1)
            elif re.search(r'\d{1,3}\.\d{1,3}\.\d{1,3}\.\d{1,3}[-]\d{1,3}', y):
                range_list = re.split(r'[-.]', y)
                y1 = 'src-addr ' + str(range_list[0]) + '.' + str(range_list[1]) + '.' + str(
                    range_list[2]) + '.' + '[' + str(range_list[3]) + '-' + str(range_list[4]) + ']'
                srcadd.append(y1)
            else:
                y1 = 'src-addr ' + y + '/32'
                srcadd.append(y1)
        x4 = x[4]
        dst_list = x4.split(' ')
        for w in dst_list:
            if re.search(r'(\d{1,3}\.\d{1,3}\.\d{1,3}\.\d)[/][2][4]', w):
                w1 = 'dst-addr ' + w
                dstadd.append(w1)
            elif re.search(r'\d{0,3}\.\d{0,3}\.\d[0]{1}\.\d[/16]{1}', w):
                w1 = 'dst-addr ' + w
                dstadd.append(w1)
            elif re.search(r'\d{1,3}\.\d{1,3}\.\d{1,3}\.\d{1,3}[-]\d{1,3}', w):
                range_list = re.split(r'[-.]', w)
                w1 = 'dst-addr ' + str(range_list[0]) + '.' + str(range_list[1]) + '.' + str(
                    range_list[2]) + '.' + '[' + str(range_list[3]) + '-' + str(range_list[4]) + ']'
                dstadd.append(w1)
            else:
                w1 = 'dst-addr ' + w + '/32'
                dstadd.append(w1)
        x6 = x[6]
        x6_str = str(x6)
        ser_list = x6_str.split(' ')
        for z in ser_list:
            z1 = 'service ' + str(x[5]).upper() + z
            serv.append(z1)
        policy_cmd.extend(
            (rule,
             act,
             szone,
             dzone,
             ('\n'.join(srcadd)),
                ('\n'.join(dstadd)),
                ('\n'.join(serv)),
                exit))
        policy_cmds = '\n'.join(policy_cmd)
    return policy_cmds


if __name__ == '__main__':
    print(read_excel('/usr/local/python3.9/path/input_files/policy_tmp.xlsx'))
    # print(create_sadd('/usr/local/python3.9/path/input_files/policy_tmp.xlsx'))
    # print(create_dadd('/usr/local/python3.9/path/input_files/policy_tmp.xlsx'))
    # print(create_service('/usr/local/python3.9/path/input_files/policy_tmp.xlsx'))
    # print(create_policy('/usr/local/python3.9/path/input_files/policy_tmp.xlsx'))
