#!/usr/bin/python3.9
# -*- coding=utf-8 -*-
# 本脚由孙艳龙编写，用于Python学习及网络自动化脚本，如有问题或技术交流请与本人联系！
# mail:sunyanlong@bris.cn

import openpyxl
import re


def read_excel(file_name):  # 读取excel数据保存为字典
    wb = openpyxl.load_workbook(file_name)
    sheet = wb.get_sheet_by_name('Sheet1')
    row = sheet.max_row
    column = sheet.max_column
    policy_list = {}
    for i in range(2, row + 1):
        policy_list['policy{0}'.format(i - 1)] = []
        for j in range(1, column + 1):
            vla = sheet.cell(row=i, column=j).value
            policy_list['policy{0}'.format(i - 1)].append(vla)
    return policy_list


def create_sadd(excelfile):  # 创建源地址模块
    sadd_dict = read_excel(excelfile)
    sadd_cmd = ['config firewall address']
    for x in sadd_dict.values():
        x2 = x[2]
        src_list = x2.split(' ')
        for y in src_list:
            if re.search(r'(\d{1,3}\.\d{1,3}\.\d{1,3}\.\d)[/][2][4]', y):
                sadd_1 = 'edit ' + str(y)
                sadd_2 = 'set subnet ' + \
                         str(y).rstrip('/24') + ' 255.255.255.0'
                sadd_3 = 'next'
                sadd_cmd.extend((sadd_1, sadd_2, sadd_3))
            elif re.search(r'\d{1,3}\.\d{1,3}\.[0]\.[0][/][1][6]', y):
                sadd_1 = 'edit ' + str(y)
                sadd_2 = 'set subnet ' + str(y).rstrip('/16') + ' 255.255.0.0'
                sadd_3 = 'next'
                sadd_cmd.extend((sadd_1, sadd_2, sadd_3))
            elif re.search(r'\d{1,3}\.\d{1,3}\.\d{1,3}\.\d{1,3}[-]\d{1,3}', y):
                range_list = re.split(r'[-.]', y)
                range_list1 = re.split(r'[-]', y)
                sadd_1 = 'edit ' + str(range_list[0]) + '.' + str(range_list[1]) + '.' + str(
                    range_list[2]) + '.' + '[' + str(range_list[3]) + '-' + str(range_list[4]) + ']'
                sadd_2 = 'set type iprange'
                sadd_3 = 'set start-ip ' + str(range_list1[0])
                sadd_4 = 'set end-ip ' + str(range_list[0]) + '.' + str(
                    range_list[1]) + '.' + str(range_list[2]) + '.' + str(range_list[4])
                sadd_5 = 'next'
                sadd_cmd.extend((sadd_1, sadd_2, sadd_3, sadd_4, sadd_5))
            else:
                sadd_1 = 'edit ' + str(y) + '/32'
                sadd_2 = 'set subnet ' + str(y) + ' 255.255.255.255'
                sadd_3 = 'next'
                sadd_cmd.extend((sadd_1, sadd_2, sadd_3))
    sadd_cmd.append('end\n')
    sadd_cmds = '\n'.join(sadd_cmd)
    return sadd_cmds


def create_dadd(excelfile):  # 创建目的地址模块
    dadd_dice = read_excel(excelfile)
    dadd_cmd = ['config firewall address']
    for x in dadd_dice.values():
        x4 = x[4]
        dst_list = x4.split(' ')
        for y in dst_list:
            if re.search(r'(\d{1,3}\.\d{1,3}\.\d{1,3}\.\d)[/][2][4]', y):
                dadd_name = 'edit ' + str(y)
                dadd_add = 'set subnet ' + \
                           str(y).rstrip('/24') + ' 255.255.255.0'
                dadd_exit = 'next'
                dadd_cmd.extend((dadd_name, dadd_add, dadd_exit))
            elif re.search(r'\d{1,3}\.\d{1,3}\.[0]\.[0][/][1][6]', y):
                dadd_name = 'edit ' + str(y)
                dadd_add = 'set subnet ' + \
                           str(y).rstrip('/16') + ' 255.255.0.0'
                dadd_exit = 'next'
                dadd_cmd.extend((dadd_name, dadd_add, dadd_exit))
            elif re.search(r'\d{1,3}\.\d{1,3}\.\d{1,3}\.\d{1,3}[-]\d{1,3}', y):
                range_list = re.split(r'[-.]', y)
                range_list1 = re.split(r'[-]', y)
                dadd_1 = 'edit ' + str(range_list[0]) + '.' + str(range_list[1]) + '.' + str(
                    range_list[2]) + '.' + '[' + str(range_list[3]) + '-' + str(range_list[4]) + ']'
                dadd_2 = 'set type iprange'
                dadd_3 = 'set start-ip ' + str(range_list1[0])
                dadd_4 = 'set end-ip ' + str(range_list[0]) + '.' + str(
                    range_list[1]) + '.' + str(range_list[2]) + '.' + str(range_list[4])
                dadd_5 = 'next'
                dadd_cmd.extend((dadd_1, dadd_2, dadd_3, dadd_4, dadd_5))
            else:
                dadd_name = 'edit ' + str(y) + '/32'
                dadd_add = 'set subnet ' + str(y) + ' 255.255.255.255'
                dadd_exit = 'next'
                dadd_cmd.extend((dadd_name, dadd_add, dadd_exit))
    dadd_cmd.append('end\n')
    dadd_cmds = '\n'.join(dadd_cmd)
    return dadd_cmds


def create_service(excelfile):  # 创建service模块
    ser_dict = read_excel(excelfile)
    ser_cmd = ['config firewall service custom']
    for x in ser_dict.values():
        x6 = x[6]
        x6_str = str(x6)
        ser_list = x6_str.split(' ')
        for y in ser_list:
            if re.search(r'\d{1,5}[-]\d{1,5}', y):
                ser_name = 'edit ' + str(x[5]).upper() + str(y)
                ser_port = 'set ' + str(x[5]) + '-portrange ' + str(y)
                ser_exit = 'next'
                ser_cmd.extend((ser_name, ser_port, ser_exit))
            else:
                ser_name = 'edit ' + str(x[5]).upper() + str(y)
                ser_port = 'set ' + str(x[5]) + '-portrange ' + str(y)
                ser_exit = 'next'
                ser_cmd.extend((ser_name, ser_port, ser_exit))
    ser_cmd.append('end\n')
    ser_cmds = '\n'.join(ser_cmd)
    return ser_cmds


def create_policy(excelfile):  # 创建策略模块
    policy_dict = read_excel(excelfile)
    policy_cmd = ['config firewall policy']
    policy_cmds = []
    for x in policy_dict.values():
        number = 'edit 0'
        src_zone = 'set srcintf ' + str(x[1])
        dst_zone = 'set dstintf ' + str(x[3])
        src_add = ['set srcaddr ']
        dst_add = ['set dstaddr ']
        ser = ['set service ']
        scherule = 'set schedule always'
        status = 'set status enable'
        next = 'next'
        x2 = x[2]
        src_list = x2.split(' ')
        for y in src_list:
            if re.search(r'(\d{1,3}\.\d{1,3}\.\d{1,3}\.\d)[/][2][4]', y):
                y1 = y
                src_add.append(y1)
            elif re.search(r'\d{1,3}\.\d{1,3}\.[0]\.[0][/][1][6]', y):
                y1 = y
                src_add.append(y1)
            elif re.search(r'\d{1,3}\.\d{1,3}\.\d{1,3}\.\d{1,3}[-]\d{1,3}', y):
                range_list = re.split(r'[-.]', y)
                y1 = str(range_list[0]) + '.' + str(range_list[1]) + '.' + str(
                    range_list[2]) + '.' + '[' + str(range_list[3]) + '-' + str(range_list[4]) + ']'
                src_add.append(y1)
            else:
                y1 = y + '/32'
                src_add.append(y1)
        x4 = x[4]
        dst_list = x4.split(' ')
        for w in dst_list:
            if re.search(r'(\d{1,3}\.\d{1,3}\.\d{1,3}\.\d)[/][2][4]', w):
                w1 = w
                dst_add.append(w1)
            elif re.search(r'\d{1,3}\.\d{1,3}\.[0]\.[0][/][1][6]', w):
                w1 = w
                dst_add.append(w1)
            elif re.search(r'\d{1,3}\.\d{1,3}\.\d{1,3}\.\d{1,3}[-]\d{1,3}', w):
                range_list = re.split(r'[-.]', w)
                w1 = str(range_list[0]) + '.' + str(range_list[1]) + '.' + str(
                    range_list[2]) + '.' + '[' + str(range_list[3]) + '-' + str(range_list[4]) + ']'
                dst_add.append(w1)
            else:
                w1 = w + '/32'
                dst_add.append(w1)
        x6 = x[6]
        x6_str = str(x6)
        ser_list = x6_str.split(' ')
        for z in ser_list:
            z1 = str(x[5]).upper() + z
            ser.append(z1)
        policy_cmd.extend((number, src_zone, dst_zone, (' '.join(src_add)), (' '.join(dst_add)), (' '.join(ser)),
                           scherule, status, next))
        policy_cmds = '\n'.join(policy_cmd)
    return policy_cmds


if __name__ == "__main__":
    print(read_excel('/usr/local/python3.9/path/input_files/policy_tmp2.xlsx'))  # 测试模块
