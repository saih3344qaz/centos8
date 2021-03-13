#!/usr/bin/python3.9
# -*- coding=utf-8 -*-
# 本脚由孙艳龙编写，用于Python学习及网络自动化脚本，如有问题或技术交流请与本人联系！
# mail:sunyanlong@bris.cn

from netmiko import ConnectHandler
import paramiko
import openpyxl
import os
import time
import ftplib
import re

now = time.strftime(
    '%Y-%m-%d_%H%M%S',
    time.localtime(
        time.time()))  # 定义时间格式，备份文件时显示日期时间用到
date = time.strftime('%Y-%m-%d', time.localtime(time.time()))  # 定义日期格式，创建文件夹用到


# 读取Excel文件并格式化，将各个设备的信息存入字典。
def read_excel(file_name):
    wb = openpyxl.load_workbook(file_name)
    sheet = wb.get_sheet_by_name('Sheet1')
    row = sheet.max_row
    column = sheet.max_column
    device_list = {}
    for i in range(2, row + 1):
        device_list['device{0}'.format(i - 1)] = []
        for j in range(1, column + 1):
            vla = sheet.cell(row=i, column=j).value
            device_list['device{0}'.format(i - 1)].append(vla)
    return device_list


# 执行命令并输出的函数
def exe_command(net_connect, cmd):
    print('正在执行命令：', cmd)
    # send_command_timing 沿通道发送命令，返回输出（基于时序）。就是紧接着回显后边执行。
    result = net_connect.send_command_timing(cmd)
    # send_command 向下发送命令，返回输出（基于模式）。就是在回显的下一行刷入命令。
    # result = net_connect.send_command(cmd)
    return result


#  设备连接方式及执行命令定义
def connect(device, backup_file_name, ty):
    print('\n正在连接:{0}'.format(device['host']))
    net_connect = ConnectHandler(**device)
    net_connect.enable()  # 输入启用
    save = open(backup_file_name, 'w')
    if ty == 'fortinet':  # 读取命令文件，存放在当前目录下
        for i in open(
                '/usr/local/python3.9/path/input_files/command_backup.txt',
                'r'):
            cmd = i.replace('\n', ' ')
            result = exe_command(net_connect, cmd)
            save.write(result)
        save.close()
        net_connect.disconnect()
    elif ty == 'cisco_ios':
        for i in open(
                '/usr/local/python3.9/path/input_files/command_backup_cisco_ios.txt',
                'r'):
            cmd = i.replace('\n', ' ')
            result = exe_command(net_connect, cmd)
            save.write(result)
        save.close()
        net_connect.disconnect()
    elif ty == 'h3c':
        for i in open('/usr/local/python3.9/path/input_files/xxx.txt', 'r'):
            cmd = i.replace('\n', ' ')
            result = exe_command(net_connect, cmd)
            save.write(result)
        save.close()
        net_connect.disconnect()
    elif ty == 'huawei':
        for i in open('/usr/local/python3.9/path/input_files/xxx.txt', 'r'):
            cmd = i.replace('\n', ' ')
            result = exe_command(net_connect, cmd)
            save.write(result)
        save.close()
        net_connect.disconnect()
        """
        根据需求增加elif添加其它设备命令
        """
    else:
        pass


#  F5是备份方式不是屏幕抓取单独定义连接方式
def f5_connect(device, hostname):
    print('\n正在连接LTM:{0}'.format(device['host']))
    net_connect = ConnectHandler(**device)
    net_connect.enable()  # 输入启用
    print(f'创建ucs文件:save sys ucs {hostname} no-private-key')
    net_connect.send_command(f'save sys ucs {hostname} no-private-key')
    time.sleep(5)
    print(f'创建ucs文件:save sys ucs {hostname} no-private-key 完成...')
    net_connect.disconnect()


#  如下是FTP上传程序
def putfile(file, rdir, ldir, verbose=True):
    hostname = '10.10.10.10'
    username = 'sunyanlong'
    password = 'saih3344!!!'
    if verbose:
        print('上传文件:', file)
    os.chdir(ldir)  # 切换本地工作目录
    local = open(file, 'rb')  # 读取本地文件

    remote = ftplib.FTP(hostname)  # 连接站点
    remote.encoding = 'GB18030'  # 使用中文编码
    remote.login(username, password)  # 输入用户名和密码进行登录
    try:
        remote.mkd('/ftp_test/' + str(date) + '/')  # 创建FTP服务器指定目录下日期文件夹
    except Exception:
        pass
    remote.cwd(rdir)  # 切换FTP目录
    remote.storbinary('STOR ' + file, local, 1024)  # 上传文件
    remote.quit()  # 退出会话
    local.close()  # 关闭本地文件
    if verbose:
        print('上传文件:' + file + ' 结束！')


def filename():
    # 获取指定路径下的所有文件将文件名存放到列表
    f1 = os.getcwd() + '/network_config/' + str(date) + '/'
    f2 = os.listdir(f1)
    filename = []
    for i in f2:
        # 正则找出所有结尾为txt/conf/DAT的文件，然后添加到filename列表里
        if ''.join(re.findall('.tx(t$)', i)) == 't':
            filename.append(i)
        elif ''.join(re.findall('.con(f$)', i)) == 'f':
            filename.append(i)
        elif ''.join(re.findall('.DA(T$)', i)) == 'T':
            filename.append(i)
        elif ''.join(re.findall('.uc(s$)', i)) == 's':
            filename.append(i)
        else:
            pass
    return filename


def ftpput_main(rdir, ldir):  # 批量上传指定目录文件到FTP
    print('正在连接FTP服务器...')
    for k in filename():
        putfile(k, rdir, ldir)
        time.sleep(1)
    print('上传FTP服务器完成...')


#  如下是SFTP下载配置文件程序
def sftp_downLoadFile(sftp, localDir, remoteDir):  # 下载单个文件
    file_handler = open(localDir, 'wb')
    print(file_handler)
    sftp.get(remoteDir, localDir)  # 下载目录中文件
    file_handler.close()
    return True


def sftp_downloadFileTree(sftp, localDir, remoteDir):
    if remoteDir.find(".") == -1:  # 判断远程目录参数是否是目录，前提是远程的文件名中都包含扩展名，否则此方法不可用
        for file in sftp.listdir(remoteDir):
            remoteDirTmp = os.path.join(remoteDir, file)
            localDirTmp = os.path.join(localDir, file)
            sftp_downloadFileTree(sftp, localDirTmp, remoteDirTmp)
    else:
        localPath = localDir.rpartition("/")[0]
        if not os.path.exists(localPath):
            os.makedirs(localPath)
        print("download file:", remoteDir)
        try:
            sftp.get(remoteDir, localDir)
        except Exception as e:
            print('download exception:', e)


def sftp_main(filenames, sftp_device):
    host = sftp_device  # sftp主机
    port = 22  # 端口
    username = 'root'  # sftp用户名
    password = 'default'
    localDir = '/usr/local/python3.9/network_config/' + \
               str(date) + '/' + filenames  # 本地文件或目录
    remoteDir = '/var/local/ucs/' + filenames  # 远程文件或目录
    sf = paramiko.Transport((host, port))
    sf.connect(username=username, password=password)
    sftp = paramiko.SFTPClient.from_transport(sf)
    print('下载ucs文件:{}'.format(filenames))
    sftp_downloadFileTree(sftp, localDir, remoteDir)
    print('下载ucs文件:{}完成...'.format(filenames))
    sf.close()


# 设备类型字典库
def H3c(ip):
    h3c = {
        'device_type': 'hp_comware',
        'host': ip,
        'username': 'admin',
        'password': 'admin',
    }
    return h3c


def Huawei(ip):
    huawei = {
        'device_type': 'huawei',
        'host': ip,
        'username': 'admin',
        'password': 'admin',
    }
    return huawei


def cisco_ios(ip):
    cisco_ios = {
        'device_type': 'cisco_ios',
        'host': ip,
        'username': 'admin',
        'password': 'admin@123',
        # 'secret': 'admin@123',               # 用户权限高可不用此条
    }
    return cisco_ios


def Fg(ip):
    fg = {
        'device_type': 'fortinet',
        'host': ip,
        'username': 'admin',
        'password': 'admin',
    }
    return fg


def F5(ip):
    f5 = {
        'device_type': 'f5_tmsh',
        'host': ip,
        'username': 'root',
        'password': 'default',
    }
    return f5


# 主函数，主要用于设备类型的判断，配置备份文件名的构造
def main_backup(ip_file):
    device_list = read_excel(ip_file)
    try:
        os.mkdir('/usr/local/python3.9/network_config/' + date)  # 创建配置目录下当日文件夹
    except Exception:
        pass
    for device_name in device_list.keys():
        try:
            if device_list[device_name][1] == 'fortinet':
                backup_dir = 'network_config/' + str(date) + '/'  # 配置存放文件夹
                device = Fg(device_list[device_name][0])
                backup_file_name = backup_dir + \
                                   device_list[device_name][0] + '_' + device_list[device_name][
                    2] + '_' + str(now) + '.conf'
                connect(device, backup_file_name, ty='fortinet')
            elif device_list[device_name][1] == 'cisco_ios':
                backup_dir = 'network_config/' + str(date) + '/'  # 配置存放文件夹
                device = cisco_ios(device_list[device_name][0])
                backup_file_name = backup_dir + \
                                   device_list[device_name][2] + '_' + str(now) + '.txt'
                connect(device, backup_file_name, ty='cisco_ios')
            elif device_list[device_name][1] == 'h3c':
                backup_dir = 'network_config/' + str(date) + '/'  # 配置存放文件夹
                device = H3c(device_list[device_name][0])
                backup_file_name = backup_dir + \
                                   device_list[device_name][2] + '_' + str(now) + '.txt'
                connect(device, backup_file_name, ty='h3c')
            elif device_list[device_name][1] == 'huawei':
                backup_dir = 'network_config/' + str(date) + '/'  # 配置存放文件夹
                device = Huawei(device_list[device_name][0])
                backup_file_name = backup_dir + \
                                   device_list[device_name][2] + '_' + str(now) + '.txt'
                connect(device, backup_file_name, ty='huawei')
            elif device_list[device_name][1] == 'f5':
                device = F5(device_list[device_name][0])
                hostname = device_list[device_name][0] + '_' + \
                           device_list[device_name][2] + '_' + str(now)
                f5_connect(device, hostname)
                time.sleep(2)
                sftp_device = device_list[device_name][0]
                sftp_main((device_list[device_name][0] +
                           '_' +
                           device_list[device_name][2] +
                           '_' +
                           str(now) +
                           '.ucs'), sftp_device)
                """
                根据需求增加elif添加其它设备类型
                """
            else:
                print('未定义设备！')
                pass
        except Exception as e:
            print('连接超时：', e)
            pass
        time.sleep(1)
    print('备份任务已结束...\n')
    ftpput_main(
        rdir='/ftp_test/' +
             str(date) +
             '/',
        ldir='/usr/local/python3.9/network_config/' +
             str(date) +
             '/')


if __name__ == '__main__':
    main_backup('/usr/local/python3.9/path/input_files/device_list_all.xlsx')
