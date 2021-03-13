#!/usr/bin/env python3
# -*- coding=utf-8 -*-
# 本脚由孙艳龙编写，用于Python学习及网络自动化脚本，如有问题或技术交流请与本人联系！
# mail:sunyanlong@bris.cn

import openpyxl
import paramiko
import time
import os
import ftplib
import re

now = time.strftime(
    '%Y-%m-%d %H%M%S',
    time.localtime(
        time.time()))  # 定义时间格式，备份文件用到
date = time.strftime('%Y-%m-%d', time.localtime(time.time()))  # 定义时间格式，备份文件用到


# 读取Excel文件并格式化，将各个设备的信息存入字典。
def read_excel(file_name):
    wb = openpyxl.load_workbook(file_name)
    sheet = wb.get_sheet_by_name('Sheet1')
    row = sheet.max_row
    column = sheet.max_column
    device_list = {}
    for i in range(2, row + 1):
        device_list['device{0}'.format(i - 1)] = []
        for j in range(1, column + 1):
            vla = sheet.cell(row=i, column=j).value
            device_list['device{0}'.format(i - 1)].append(vla)
    return device_list


def ssh_connect(device, backup_file_name):
    username = 'hillstone'
    password = 'admin@123'
    ssh = paramiko.SSHClient()  # 创建SSH Client
    ssh.load_system_host_keys()  # 加载系统SSH密钥
    ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())  # 添加新的SSH密钥
    # print('正在连接{0}\n'.format(device['host']))
    ssh.connect(
        device,
        port=22,
        username=username,
        password=password,
        timeout=5,
        compress=True)  # SSH连接
    chan = ssh.invoke_shell()  # 激活交互式shell
    time.sleep(2)
    # x = chan.recv(2048).decode()  # 接收回显信息
    # print('\n开始保存......')
    for cmd in open(
        '/usr/local/python3.9/path/input_files/command_backup_ss.txt',
            'r'):
        print('执行命令:' + str(cmd).strip('\n'))
        chan.send(cmd.encode())  # 执行命令，注意字串都需要编码为二进制字串
        chan.send(b'\n')  # 一定要注意输入回车
        time.sleep(5)  # 由于有些回显可能过长，所以可以考虑等待更长一些时间
        x = chan.recv(4096000).decode(
            errors='ignore')  # 读取回显，有些回想可能过长，请把接收缓存调大
        save = open(backup_file_name, 'w')
        save.write(x)
        save.close()
    print('\n保存完成......')
    chan.close()  # 退出交互式shell
    ssh.close()  # 退出ssh会话


def putfile(file, rdir, ldir, verbose=True):
    hostname = '10.10.10.10'
    username = 'sunyanlong'
    password = 'saih3344!!!'
    if verbose:
        print('上传文件:', file)
    os.chdir(ldir)  # 切换本地工作目录
    local = open(file, 'rb')  # 读取本地文件

    remote = ftplib.FTP(hostname)  # 连接站点
    remote.encoding = 'GB18030'  # 使用中文编码
    remote.login(username, password)  # 输入用户名和密码进行登录
    try:
        remote.mkd('/ftp_test/' + str(date) + '/')
    except Exception:
        pass
    remote.cwd(rdir)  # 切换FTP目录
    remote.storbinary('STOR ' + file, local, 1024)  # 上传文件
    remote.quit()  # 退出会话
    local.close()  # 关闭本地文件
    if verbose:
        print('上传文件:' + file + ' 结束！')


def filename():
    # 获取指定路径下的所有文件将文件名存放到列表
    f1 = os.getcwd() + '/hillstone_config/' + str(date) + '/'
    f2 = os.listdir(f1)
    filename = []
    for i in f2:
        # 正则找出所有结尾为txt/DAT的文件，然后添加到filename列表里
        if ''.join(re.findall('.tx(t$)', i)) == 't':
            filename.append(i)
        elif ''.join(re.findall('.DA(T$)', i)) == 'T':
            filename.append(i)
        else:
            pass
    return filename


def ftp_main(rdir, ldir):
    print('正在连接FTP服务器...')
    for k in filename():
        putfile(k, rdir, ldir)
        time.sleep(1)
    print('上传FTP服务器完成...')
    return


def main(ip_file):
    device_list = read_excel(ip_file)
    try:
        os.mkdir('/usr/local/python3.9/hillstone_config/' + date)  # 创建日期目录
    except Exception:
        pass
    for device_name in device_list.values():
        try:
            backup_dir = 'hillstone_config/' + str(date) + '/'  # 配置文件夹
            device = device_name[0]
            backup_file_name = backup_dir + \
                device_name[0] + '_' + device_name[2] + '_' + str(now) + '.DAT'
            print('\n正在连接:{0}'.format(device_name[0]))
            ssh_connect(device, backup_file_name)
        except Exception as e:
            print('连接超时：', e)
            pass
        time.sleep(1)
    print('备份任务已结束...\n')
    ftp_main(
        rdir='/ftp_test/' +
        str(date) +
        '/',
        ldir='/usr/local/python3.9/hillstone_config/' +
        str(date) +
        '/')


if __name__ == '__main__':
    main('/usr/local/python3.9/path/input_files/device_list_ss.xlsx')
